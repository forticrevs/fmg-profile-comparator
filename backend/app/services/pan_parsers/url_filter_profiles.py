"""Extract Palo Alto URL-filtering profiles into Excel workbooks.

Produces two workbooks in one run:
  - `url-filter-profiles.xlsx` — one sheet per profile, two columns
    (alert categories, block categories). Sheet names are truncated to
    Excel's 31-char limit.
  - `url-filter-all-categories.xlsx` — a single "AllCategories" sheet
    listing every unique URL category seen across all profiles, with a
    semicolon-joined list of the profiles that reference it.

Ported from `xml_url_filter_to_excel.py`. Uses openpyxl directly (the
original used pandas, which we don't need as a dep). The original
script's third workbook — a hard-coded filter for a specific set of
customer profile names — is intentionally dropped; if per-profile
filtering is useful later it belongs as a UI control, not a compile-time
constant.
"""

from __future__ import annotations

import io
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.pan_parsers import register

URLF_XPATH = ".//profiles/url-filtering/entry"


def _members(parent: Any, tag: str) -> list[str]:
    el = parent.find(tag)
    if el is None:
        return []
    return [m.text for m in el.findall("member") if m.text]


def _safe_sheet_name(raw: str, used: set[str]) -> str:
    """Excel sheet names max out at 31 chars and must be unique within a workbook."""
    name = (raw or "unnamed")[:31]
    base = name
    counter = 1
    while name in used:
        suffix = f"~{counter}"
        name = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(name)
    return name


def parse(root: Any) -> dict[str, bytes]:
    profiles: list[dict[str, Any]] = []
    cat_to_profiles: dict[str, set[str]] = defaultdict(set)

    for entry in root.findall(URLF_XPATH):
        name = entry.get("name") or ""
        alert = _members(entry, "alert")
        block = _members(entry, "block")
        profiles.append({"name": name, "alert": alert, "block": block})
        for cat in alert + block:
            cat_to_profiles[cat].add(name)

    # ── Workbook 1: one sheet per profile ────────────────────────────
    wb_profiles = Workbook()
    default = wb_profiles.active
    if default is not None:
        wb_profiles.remove(default)

    used_names: set[str] = set()
    for prof in profiles:
        sheet_name = _safe_sheet_name(prof["name"], used_names)
        ws = wb_profiles.create_sheet(sheet_name)
        ws.append(["Alert Categories", "Block Categories"])
        max_len = max(len(prof["alert"]), len(prof["block"]))
        for i in range(max_len):
            ws.append([
                prof["alert"][i] if i < len(prof["alert"]) else "",
                prof["block"][i] if i < len(prof["block"]) else "",
            ])
        for col in (1, 2):
            ws.column_dimensions[get_column_letter(col)].width = 42

    # If no profiles were found, openpyxl refuses to save a 0-sheet
    # workbook — give it a placeholder.
    if not wb_profiles.sheetnames:
        wb_profiles.create_sheet("empty").append(["No URL-filtering profiles found"])

    buf1 = io.BytesIO()
    wb_profiles.save(buf1)

    # ── Workbook 2: all categories with referencing profiles ─────────
    wb_allcats = Workbook()
    ws = wb_allcats.active
    if ws is None:
        ws = wb_allcats.create_sheet("AllCategories")
    else:
        ws.title = "AllCategories"

    ws.append(["URL Category", "Profiles"])
    for cat in sorted(cat_to_profiles.keys()):
        ws.append([cat, "; ".join(sorted(cat_to_profiles[cat]))])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 80

    buf2 = io.BytesIO()
    wb_allcats.save(buf2)

    return {
        "url-filter-profiles.xlsx": buf1.getvalue(),
        "url-filter-all-categories.xlsx": buf2.getvalue(),
    }


register(
    parser_id="url_filter_profiles",
    label="URL Filter Profiles",
    description="URL-filtering profiles (per-profile alert/block sheets) plus a unified categories workbook listing every category and the profiles that reference it",
    parse=parse,
)
